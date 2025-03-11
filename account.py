# -*- coding: utf-8 -*-
import tkinter
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import re
import datetime
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font

def calculate_time_difference(input_str):
    # 使用正則表達式提取時間，支持 '~' 和 '～' 作為時間分隔符
    time_pattern = r"(\d{1,2}:\d{2})[~～](\d{1,2}:\d{2})"
    match = re.search(time_pattern, input_str)

    if match:
        # 取得開始和結束時間
        start_time_str = match.group(1)
        end_time_str = match.group(2)

        # 轉換為 datetime 物件
        start_time = datetime.datetime.strptime(start_time_str, "%H:%M")
        end_time = datetime.datetime.strptime(end_time_str, "%H:%M")

        # 如果結束時間早於開始時間，表示跨天，將結束時間加一天
        if end_time < start_time:
            end_time += datetime.timedelta(days=1)

        # 計算時間差
        time_difference = end_time - start_time
        hours = time_difference.seconds // 3600
        minutes = (time_difference.seconds // 60) % 60

        # 根據分鐘數決定返回時數
        if minutes == 0:
            return hours
        elif minutes == 30:
            return hours + 0.5
    
    # 如果沒有找到符合時間格式的字串，則返回 0
    return 0

class MainSerial:
    def __init__(self):
        # 設定程式名稱
        self.mainwin = tkinter.Tk()
        self.mainwin.title("記帳程式") 
        self.mainwin.geometry("400x200")

        # 標籤
        self.label1 = tkinter.Label(self.mainwin, text="讀取檔案:", font=("標楷體", 15))
        self.label1.place(x=5, y=5)

        # 按鈕樣式
        s = ttk.Style()
        s.configure('my.TButton', font=('標楷體', 15))

        # 創建「選擇檔案」按鈕
        self.upload_btn = ttk.Button(self.mainwin, text="選擇 Excel 檔案", style='my.TButton', command=self.upload_file)
        self.upload_btn.place(x=105, y=5)

    #============================================================================================

            
    def upload_file(self):
        labels = ["日期", "時間/客戶","時數","時薪","時薪","油資","獎勵金","誤餐","合計"]
        file_path = filedialog.askopenfilename(title="選擇檔案", filetypes=[("Excel 檔案", "*.xlsx;*.xls;*.csv")])
        
        if not os.path.exists('./員工班表'):
            os.mkdir('./員工班表')
        
        if not file_path:  # 確保用戶有選擇檔案
            return

        try:
            df = pd.read_excel(file_path, header=[0, 0])  # 讀取 Excel
            output_data = []
            flage1 = 0
            flage2 = 0
            
            name_list = []
            for (day, date) in df.columns[1:]:
                for i, row in df.iterrows():
                    name = row[df.columns[0]]  # 取得名字
                    if (pd.isna(name) or name == 'nan'):
                        continue
                    name_list.append(name)
                    
            unique_names = list(set(name_list))
            unique_names.sort(key=name_list.index)
            
            for n in range(len(unique_names)):
                # 嘗試開啟現有 Excel，否則建立新檔案
                save_path = "./員工班表/"+str(unique_names[n])+"_班表.xlsx"
                time_count = 0
                money_count = 0
                total_count = 0
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(labels)  # 若是新檔案，添加標題
                
                '''
                # 讓文字置中
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                '''
            
            # 逐行處理 Excel 資料
                for (day, date) in df.columns[1:]:
                    for i, row in df.iterrows():
                        name = row[df.columns[0]]  # 取得名字

                        if name == unique_names[n]:  # 只處理特定使用者
                        
                            schedule1 = row[(day, date)]
                            schedule2 = None
                            
                            if i + 1 < len(df): 
                                next_row = df.iloc[i + 1]
                                schedule2 = next_row[(day, date)]
                            
                            # 判斷是否有數據
                            if pd.notna(schedule1):  # 跳過空白欄位
                                flage1 = 1
                            if pd.notna(schedule2):  # 跳過空白欄位
                                flage2 = 1
                            
                        # 解析日期
                        value = row[(day, date)]
                        if (pd.isna(value) or value == 'nan'):
                            continue

                        if isinstance(value, (pd.Timestamp, datetime.datetime)):
                            date_str = value.strftime("%m月%d日")
                            #print(date_str)

                        # 準備寫入 Excel
                        new_entries = []
                        if (flage1 == 1) and (flage2 == 0):
                            flage1 = 0
                            hours = calculate_time_difference(schedule1)
                            new_entries.append([date_str, str(schedule1), hours, 270, hours*270,"","","",hours*270])
                            time_count += hours
                            money_count += (hours*270)
                            total_count += (hours*270)
                            
                        elif (flage1 == 0) and (flage2 == 1):
                            flage2 = 0
                            hours = calculate_time_difference(schedule2)
                            new_entries.append([date_str, str(schedule2), hours, 270, hours*270,"","","",hours*270])
                            time_count += hours
                            money_count += (hours*270)
                            total_count += (hours*270)
                        elif (flage1 == 1) and (flage2 == 1):
                            flage1 = 0
                            flage2 = 0
                            hours = calculate_time_difference(schedule1)
                            new_entries.append([date_str, str(schedule1), hours, 270, hours*270,"","","",hours*270]) 
                            time_count += hours
                            money_count += (hours*270)
                            total_count += (hours*270)
                            hours = calculate_time_difference(schedule2)
                            new_entries.append([date_str, str(schedule2), hours, 270, hours*270,"","","",hours*270])
                            time_count += hours
                            money_count += (hours*270)
                            total_count += (hours*270)
                        
                        # 調整excel格子大小
                        for col_num in range(1, sheet.max_column + 1):
                            col_letter = get_column_letter(col_num)
                            #print(col_letter)
                            if col_letter == 'B':
                                sheet.column_dimensions[col_letter].width = 20

                        # 寫入 Excel
                        for entry in new_entries:
                            sheet.append(entry)
                
                sheet.append(["","","","","","","","",""])
                sheet.append(["", "總計", time_count,"", money_count,"","","", total_count])

                # 儲存 Excel
                workbook.save(save_path)
                workbook.close()
                messagebox.showinfo("成功", "數據已成功寫入 "+str(unique_names[n])+"_班表.xlsx")

        except Exception as e:
            messagebox.showerror("錯誤", f"讀取 Excel 失敗：{e}")

    #============================================================================================
    
    def show(self):
        self.mainwin.mainloop()

if __name__ == "__main__":
    my_ser1 = MainSerial()
    my_ser1.show()